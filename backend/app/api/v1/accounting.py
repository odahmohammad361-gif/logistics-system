import csv
import io
import json
import os
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, Response
from sqlalchemy import func, or_
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.dependencies import get_current_user, require_role
from app.database import get_db
from app.models.accounting import (
    AccountingAttachment,
    AccountingDirection,
    AccountingEntry,
    AccountingStatus,
    BankLineMatchStatus,
    BankStatementImport,
    BankStatementLine,
)
from app.models.booking import Booking
from app.models.branch import Branch
from app.models.clearance_agent import ClearanceAgent
from app.models.client import Client
from app.models.invoice import Invoice
from app.models.shipping_agent import ShippingAgent
from app.models.supplier import Supplier
from app.models.user import User, UserRole
from app.schemas.accounting import (
    AccountingAttachmentResponse,
    AccountingEntryCreate,
    AccountingEntryListResponse,
    AccountingEntryResponse,
    AccountingEntryUpdate,
    AccountingReportSummaryResponse,
    AccountingCategoryTotal,
    AccountingTaxAlert,
    AccountingSummaryResponse,
    BankStatementImportListResponse,
    BankStatementLineListResponse,
    BankStatementUploadResponse,
)

router = APIRouter()

UPLOAD_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
    "uploads",
    "accounting",
)
os.makedirs(UPLOAD_DIR, exist_ok=True)

DOCUMENT_TYPES = {"receipt", "invoice", "bank_proof", "salary", "expense", "other"}

DATE_KEYS = {
    "date", "transaction date", "posting date", "value date", "تاريخ", "التاريخ", "تاريخ الحركة", "تاريخ العملية",
}
DESCRIPTION_KEYS = {
    "description", "details", "detail", "narrative", "memo", "beneficiary", "البيان", "الوصف", "التفاصيل", "شرح",
}
REFERENCE_KEYS = {
    "reference", "ref", "reference no", "reference number", "transaction id", "رقم المرجع", "مرجع", "رقم العملية",
}
AMOUNT_KEYS = {"amount", "transaction amount", "المبلغ", "قيمة الحركة"}
DEBIT_KEYS = {"debit", "withdrawal", "withdrawals", "paid out", "مدين", "سحب", "خصم"}
CREDIT_KEYS = {"credit", "deposit", "deposits", "paid in", "دائن", "ايداع", "إيداع"}
BALANCE_KEYS = {"balance", "running balance", "الرصيد", "الرصيد المتاح"}


def _generate_entry_number(db: Session, direction: AccountingDirection, entry_date: date) -> str:
    prefix = "REC" if direction == AccountingDirection.MONEY_IN else "PAY"
    pattern = f"{prefix}-{entry_date.year}-%"
    count = db.query(AccountingEntry).filter(AccountingEntry.entry_number.like(pattern)).count()
    return f"{prefix}-{entry_date.year}-{str(count + 1).zfill(5)}"


def _get_entry(db: Session, entry_id: int) -> AccountingEntry:
    entry = db.query(AccountingEntry).filter(AccountingEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Accounting entry not found")
    return entry


def _ensure_linked_records_exist(db: Session, payload: AccountingEntryCreate | AccountingEntryUpdate) -> None:
    checks = (
        ("client_id", Client, "Client"),
        ("invoice_id", Invoice, "Invoice"),
        ("booking_id", Booking, "Container"),
        ("shipping_agent_id", ShippingAgent, "Shipping agent"),
        ("clearance_agent_id", ClearanceAgent, "Clearance agent"),
        ("supplier_id", Supplier, "Supplier"),
        ("branch_id", Branch, "Branch"),
    )
    for field, model, label in checks:
        value = getattr(payload, field, None)
        if value is None:
            continue
        if not db.query(model).filter(model.id == value).first():
            raise HTTPException(404, f"{label} not found")


def _sum_direction(db: Session, direction: AccountingDirection) -> Decimal:
    total = (
        db.query(func.coalesce(func.sum(AccountingEntry.amount), 0))
        .filter(AccountingEntry.direction == direction.value, AccountingEntry.status != AccountingStatus.VOID.value)
        .scalar()
    )
    return Decimal(str(total or 0))


def _norm(value: object) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _clean_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace(",", "").replace("JOD", "").replace("USD", "").replace("$", "").replace("د.ا", "").strip("() ")
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _find_key(row: dict[str, object], keys: set[str]) -> str | None:
    normalized = {_norm(k): k for k in row.keys()}
    for key in keys:
        if key in normalized:
            return normalized[key]
    for norm_key, original in normalized.items():
        if any(key in norm_key for key in keys if len(key) > 3):
            return original
    return None


def _row_to_line(row: dict[str, object], currency: str) -> dict | None:
    date_key = _find_key(row, DATE_KEYS)
    if not date_key:
        return None
    transaction_date = _parse_date(row.get(date_key))
    if not transaction_date:
        return None

    debit_key = _find_key(row, DEBIT_KEYS)
    credit_key = _find_key(row, CREDIT_KEYS)
    amount_key = _find_key(row, AMOUNT_KEYS)
    balance_key = _find_key(row, BALANCE_KEYS)
    description_key = _find_key(row, DESCRIPTION_KEYS)
    reference_key = _find_key(row, REFERENCE_KEYS)

    amount: Decimal | None = None
    if debit_key or credit_key:
        debit = _clean_decimal(row.get(debit_key)) if debit_key else Decimal("0")
        credit = _clean_decimal(row.get(credit_key)) if credit_key else Decimal("0")
        amount = (credit or Decimal("0")) - (debit or Decimal("0"))
    elif amount_key:
        amount = _clean_decimal(row.get(amount_key))

    if amount is None or amount == 0:
        return None

    direction = AccountingDirection.MONEY_IN.value if amount > 0 else AccountingDirection.MONEY_OUT.value
    abs_amount = abs(amount).quantize(Decimal("0.01"))
    balance = _clean_decimal(row.get(balance_key)) if balance_key else None
    return {
        "transaction_date": transaction_date,
        "direction": direction,
        "amount": abs_amount,
        "currency": currency,
        "description": str(row.get(description_key) or "").strip() if description_key else None,
        "reference_no": str(row.get(reference_key) or "").strip() if reference_key else None,
        "balance": balance.quantize(Decimal("0.01")) if balance is not None else None,
        "raw_data": json.dumps({str(k): str(v) for k, v in row.items()}, ensure_ascii=False),
    }


def _parse_csv(path: str) -> list[dict[str, object]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as fp:
        sample = fp.read(4096)
        fp.seek(0)
        dialect = csv.Sniffer().sniff(sample or ",")
        return [dict(row) for row in csv.DictReader(fp, dialect=dialect)]


def _parse_xlsx(path: str) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx = 0
    for idx, row in enumerate(rows[:15]):
        normalized = {_norm(cell) for cell in row if cell is not None}
        if normalized & DATE_KEYS and (normalized & AMOUNT_KEYS or normalized & DEBIT_KEYS or normalized & CREDIT_KEYS):
            header_idx = idx
            break
    headers = [str(c or "").strip() for c in rows[header_idx]]
    parsed: list[dict[str, object]] = []
    for row in rows[header_idx + 1:]:
        parsed.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]})
    return parsed


def _parse_statement_file(path: str, currency: str) -> list[dict]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        rows = _parse_csv(path)
    elif ext in {".xlsx", ".xlsm"}:
        rows = _parse_xlsx(path)
    else:
        raise HTTPException(400, "Stage 3 supports CSV and XLSX bank statements. PDF/image OCR comes later.")
    lines = [_row_to_line(row, currency) for row in rows]
    return [line for line in lines if line is not None]


def _match_line(db: Session, line: BankStatementLine) -> None:
    candidates = (
        db.query(AccountingEntry)
        .filter(
            AccountingEntry.direction == line.direction,
            AccountingEntry.status != AccountingStatus.VOID.value,
            AccountingEntry.amount == line.amount,
        )
        .order_by(AccountingEntry.entry_date.desc(), AccountingEntry.id.desc())
        .limit(50)
        .all()
    )
    best: tuple[int, AccountingEntry, str] | None = None
    line_ref = (line.reference_no or "").lower()
    line_desc = (line.description or "").lower()

    for entry in candidates:
        score = 50
        reasons = ["same amount", "same direction"]
        delta = abs((entry.entry_date - line.transaction_date).days)
        if delta == 0:
            score += 25
            reasons.append("same date")
        elif delta <= 3:
            score += 15
            reasons.append(f"date within {delta} day(s)")
        ref = (entry.reference_no or "").lower()
        if ref and (ref in line_ref or line_ref in ref):
            score += 25
            reasons.append("reference match")
        counterparty = (entry.counterparty_name or "").lower()
        if counterparty and counterparty in line_desc:
            score += 10
            reasons.append("counterparty in description")
        if best is None or score > best[0]:
            best = (score, entry, ", ".join(reasons))

    if best and best[0] >= 90:
        line.match_status = BankLineMatchStatus.MATCHED.value
        line.matched_entry_id = best[1].id
        line.match_confidence = min(best[0], 100)
        line.match_reason = best[2]
    elif best and best[0] >= 60:
        line.match_status = BankLineMatchStatus.POSSIBLE.value
        line.matched_entry_id = best[1].id
        line.match_confidence = min(best[0], 100)
        line.match_reason = best[2]
    else:
        line.match_status = BankLineMatchStatus.UNMATCHED.value
        line.match_confidence = 0
        line.match_reason = "No accounting entry with same amount/direction found"


def _period_entries_query(
    db: Session,
    date_from: date | None,
    date_to: date | None,
    currency: str,
):
    q = db.query(AccountingEntry).filter(
        AccountingEntry.status != AccountingStatus.VOID.value,
        AccountingEntry.currency == currency.upper(),
    )
    if date_from:
        q = q.filter(AccountingEntry.entry_date >= date_from)
    if date_to:
        q = q.filter(AccountingEntry.entry_date <= date_to)
    return q


def _period_bank_lines_query(
    db: Session,
    date_from: date | None,
    date_to: date | None,
    currency: str,
):
    q = db.query(BankStatementLine).filter(BankStatementLine.currency == currency.upper())
    if date_from:
        q = q.filter(BankStatementLine.transaction_date >= date_from)
    if date_to:
        q = q.filter(BankStatementLine.transaction_date <= date_to)
    return q


def _sum_query_amount(q) -> Decimal:
    total = q.with_entities(func.coalesce(func.sum(AccountingEntry.amount), 0)).scalar()
    return Decimal(str(total or 0))


def _build_report_summary(
    db: Session,
    date_from: date | None,
    date_to: date | None,
    currency: str,
) -> AccountingReportSummaryResponse:
    cur = currency.upper()
    base = _period_entries_query(db, date_from, date_to, cur)
    money_in = _sum_query_amount(base.filter(AccountingEntry.direction == AccountingDirection.MONEY_IN.value))
    money_out = _sum_query_amount(base.filter(AccountingEntry.direction == AccountingDirection.MONEY_OUT.value))
    tax_on_income = _sum_query_amount(
        base.filter(
            AccountingEntry.direction == AccountingDirection.MONEY_IN.value,
            AccountingEntry.tax_amount.isnot(None),
        )
    )
    tax_on_expenses = _sum_query_amount(
        base.filter(
            AccountingEntry.direction == AccountingDirection.MONEY_OUT.value,
            AccountingEntry.tax_amount.isnot(None),
        )
    )
    # The sums above use entry amount for consistency with existing Decimal helper; below fixes them to tax sums.
    tax_on_income = Decimal(str(
        base.filter(AccountingEntry.direction == AccountingDirection.MONEY_IN.value)
        .with_entities(func.coalesce(func.sum(AccountingEntry.tax_amount), 0))
        .scalar() or 0
    ))
    tax_on_expenses = Decimal(str(
        base.filter(AccountingEntry.direction == AccountingDirection.MONEY_OUT.value)
        .with_entities(func.coalesce(func.sum(AccountingEntry.tax_amount), 0))
        .scalar() or 0
    ))

    missing_q = base.filter(
        AccountingEntry.direction == AccountingDirection.MONEY_OUT.value,
        AccountingEntry.has_official_tax_invoice == False,
        AccountingEntry.status != AccountingStatus.VOID.value,
    )
    missing_amount = _sum_query_amount(missing_q)
    needs_review_count = base.filter(AccountingEntry.status == AccountingStatus.NEEDS_REVIEW.value).count()
    unmatched_bank_lines = _period_bank_lines_query(db, date_from, date_to, cur).filter(
        BankStatementLine.match_status == BankLineMatchStatus.UNMATCHED.value
    ).count()

    category_rows = (
        base.with_entities(
            AccountingEntry.direction,
            AccountingEntry.category,
            func.coalesce(func.sum(AccountingEntry.amount), 0).label("amount"),
            func.count(AccountingEntry.id).label("count"),
        )
        .group_by(AccountingEntry.direction, AccountingEntry.category)
        .order_by(AccountingEntry.direction, AccountingEntry.category)
        .all()
    )

    tax_alert_entries = (
        missing_q.order_by(AccountingEntry.entry_date.desc(), AccountingEntry.id.desc())
        .limit(25)
        .all()
    )

    return AccountingReportSummaryResponse(
        date_from=date_from,
        date_to=date_to,
        currency=cur,
        money_in=money_in,
        money_out=money_out,
        net=money_in - money_out,
        tax_on_income=tax_on_income,
        tax_on_expenses=tax_on_expenses,
        tax_net=tax_on_income - tax_on_expenses,
        needs_review_count=needs_review_count,
        missing_official_invoice_count=missing_q.count(),
        missing_official_invoice_amount=missing_amount,
        unmatched_bank_lines=unmatched_bank_lines,
        category_totals=[
            AccountingCategoryTotal(
                direction=row.direction,
                category=row.category,
                amount=Decimal(str(row.amount or 0)),
                count=int(row.count or 0),
            )
            for row in category_rows
        ],
        tax_alerts=[
            AccountingTaxAlert(
                id=e.id,
                entry_number=e.entry_number,
                entry_date=e.entry_date,
                amount=e.amount,
                currency=e.currency,
                category=e.category,
                counterparty_name=e.counterparty_name,
                description=e.description,
                reason="Official tax invoice/proof is not marked as attached",
            )
            for e in tax_alert_entries
        ],
    )


@router.get("/summary", response_model=AccountingSummaryResponse)
def get_summary(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    money_in = _sum_direction(db, AccountingDirection.MONEY_IN)
    money_out = _sum_direction(db, AccountingDirection.MONEY_OUT)
    needs_review = (
        db.query(AccountingEntry)
        .filter(AccountingEntry.status == AccountingStatus.NEEDS_REVIEW.value)
        .count()
    )
    recent_count = (
        db.query(AccountingEntry)
        .filter(AccountingEntry.status != AccountingStatus.VOID.value)
        .count()
    )
    return AccountingSummaryResponse(
        money_in=money_in,
        money_out=money_out,
        balance=money_in - money_out,
        needs_review=needs_review,
        recent_count=recent_count,
    )


@router.get("/reports/summary", response_model=AccountingReportSummaryResponse)
def get_report_summary(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    currency: str = "USD",
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return _build_report_summary(db, date_from, date_to, currency)


@router.get("/reports/summary.csv")
def export_report_summary_csv(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    currency: str = "USD",
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    report = _build_report_summary(db, date_from, date_to, currency)
    fp = io.StringIO()
    writer = csv.writer(fp)
    writer.writerow(["Accounting Report", report.currency])
    writer.writerow(["Date From", report.date_from or ""])
    writer.writerow(["Date To", report.date_to or ""])
    writer.writerow([])
    writer.writerow(["Money In", report.money_in])
    writer.writerow(["Money Out", report.money_out])
    writer.writerow(["Net", report.net])
    writer.writerow(["Tax On Income", report.tax_on_income])
    writer.writerow(["Tax On Expenses", report.tax_on_expenses])
    writer.writerow(["Tax Net", report.tax_net])
    writer.writerow(["Needs Review", report.needs_review_count])
    writer.writerow(["Missing Official Invoice Count", report.missing_official_invoice_count])
    writer.writerow(["Missing Official Invoice Amount", report.missing_official_invoice_amount])
    writer.writerow(["Unmatched Bank Lines", report.unmatched_bank_lines])
    writer.writerow([])
    writer.writerow(["Direction", "Category", "Amount", "Count"])
    for row in report.category_totals:
        writer.writerow([row.direction.value, row.category, row.amount, row.count])
    writer.writerow([])
    writer.writerow(["Tax Alerts"])
    writer.writerow(["Entry", "Date", "Amount", "Category", "Counterparty", "Reason"])
    for alert in report.tax_alerts:
        writer.writerow([
            alert.entry_number,
            alert.entry_date,
            alert.amount,
            alert.category,
            alert.counterparty_name or "",
            alert.reason,
        ])
    filename = f"accounting-report-{date.today().isoformat()}.csv"
    return Response(
        content=fp.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/entries", response_model=AccountingEntryListResponse)
def list_entries(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    direction: Optional[AccountingDirection] = None,
    status_filter: Optional[AccountingStatus] = Query(None, alias="status"),
    search: str = "",
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(AccountingEntry)
    if direction:
        q = q.filter(AccountingEntry.direction == direction.value)
    if status_filter:
        q = q.filter(AccountingEntry.status == status_filter.value)
    if date_from:
        q = q.filter(AccountingEntry.entry_date >= date_from)
    if date_to:
        q = q.filter(AccountingEntry.entry_date <= date_to)
    if search:
        term = f"%{search.strip()}%"
        q = q.filter(
            or_(
                AccountingEntry.entry_number.ilike(term),
                AccountingEntry.reference_no.ilike(term),
                AccountingEntry.counterparty_name.ilike(term),
                AccountingEntry.description.ilike(term),
                AccountingEntry.category.ilike(term),
            )
        )

    total = q.count()
    results = (
        q.order_by(AccountingEntry.entry_date.desc(), AccountingEntry.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return AccountingEntryListResponse(total=total, page=page, page_size=page_size, results=results)


@router.post("/entries", response_model=AccountingEntryResponse, status_code=status.HTTP_201_CREATED)
def create_entry(
    payload: AccountingEntryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.STAFF)),
):
    _ensure_linked_records_exist(db, payload)
    data = payload.model_dump()
    data["direction"] = payload.direction.value
    data["status"] = payload.status.value
    entry = AccountingEntry(
        **data,
        entry_number=_generate_entry_number(db, payload.direction, payload.entry_date),
        created_by_id=current_user.id,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


@router.get("/entries/{entry_id}", response_model=AccountingEntryResponse)
def get_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return _get_entry(db, entry_id)


@router.patch("/entries/{entry_id}", response_model=AccountingEntryResponse)
def update_entry(
    entry_id: int,
    payload: AccountingEntryUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.STAFF)),
):
    _ensure_linked_records_exist(db, payload)
    entry = _get_entry(db, entry_id)
    update_data = payload.model_dump(exclude_unset=True)
    if payload.status is not None:
        update_data["status"] = payload.status.value
    for field, value in update_data.items():
        setattr(entry, field, value)
    db.commit()
    db.refresh(entry)
    return entry


@router.delete("/entries/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
def void_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.ADMIN)),
):
    entry = _get_entry(db, entry_id)
    entry.status = AccountingStatus.VOID.value
    db.commit()


@router.post("/entries/{entry_id}/attachments", response_model=list[AccountingAttachmentResponse])
async def upload_entry_attachments(
    entry_id: int,
    document_type: str = Query("receipt"),
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.STAFF)),
):
    doc_type = document_type.strip().lower()
    if doc_type not in DOCUMENT_TYPES:
        raise HTTPException(400, "document_type must be receipt, invoice, bank_proof, salary, expense, or other")

    entry = _get_entry(db, entry_id)
    entry_dir = os.path.join(UPLOAD_DIR, str(entry.id))
    os.makedirs(entry_dir, exist_ok=True)

    created: list[AccountingAttachment] = []
    for upload in files:
        content = await upload.read()
        ext = os.path.splitext(upload.filename or "file.bin")[1].lower() or ".bin"
        fname = f"{uuid.uuid4().hex}{ext}"
        full_path = os.path.join(entry_dir, fname)
        with open(full_path, "wb") as fp:
            fp.write(content)
        attachment = AccountingAttachment(
            entry_id=entry.id,
            document_type=doc_type,
            file_path=f"accounting/{entry.id}/{fname}",
            original_filename=upload.filename,
            content_type=upload.content_type,
            file_size=len(content),
            uploaded_by_id=current_user.id,
        )
        db.add(attachment)
        created.append(attachment)

    if entry.status == AccountingStatus.DRAFT.value:
        entry.status = AccountingStatus.NEEDS_REVIEW.value
    db.commit()
    for attachment in created:
        db.refresh(attachment)
    return created


@router.get("/entries/{entry_id}/attachments/{attachment_id}")
def serve_entry_attachment(
    entry_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    attachment = (
        db.query(AccountingAttachment)
        .filter(AccountingAttachment.id == attachment_id, AccountingAttachment.entry_id == entry_id)
        .first()
    )
    if not attachment:
        raise HTTPException(404, "Attachment not found")
    uploads_root = os.path.dirname(UPLOAD_DIR)
    full_path = os.path.join(uploads_root, attachment.file_path)
    if not os.path.isfile(full_path):
        raise HTTPException(404, "Attachment file not found on disk")
    return FileResponse(full_path, filename=attachment.original_filename)


@router.delete("/entries/{entry_id}/attachments/{attachment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_entry_attachment(
    entry_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.STAFF)),
):
    attachment = (
        db.query(AccountingAttachment)
        .filter(AccountingAttachment.id == attachment_id, AccountingAttachment.entry_id == entry_id)
        .first()
    )
    if not attachment:
        raise HTTPException(404, "Attachment not found")
    uploads_root = os.path.dirname(UPLOAD_DIR)
    full_path = os.path.join(uploads_root, attachment.file_path)
    if os.path.isfile(full_path):
        os.remove(full_path)
    db.delete(attachment)
    db.commit()


@router.get("/bank-statements", response_model=BankStatementImportListResponse)
def list_bank_statements(
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(BankStatementImport).order_by(BankStatementImport.created_at.desc())
    return BankStatementImportListResponse(total=q.count(), results=q.limit(limit).all())


@router.post("/bank-statements/upload", response_model=BankStatementUploadResponse)
async def upload_bank_statement(
    file: UploadFile = File(...),
    bank_name: Optional[str] = None,
    account_name: Optional[str] = None,
    account_no: Optional[str] = None,
    currency: str = "USD",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.STAFF)),
):
    ext = os.path.splitext(file.filename or "statement.csv")[1].lower()
    if ext not in {".csv", ".xlsx", ".xlsm"}:
        raise HTTPException(400, "Upload CSV or XLSX bank statement files for this stage")

    statement_dir = os.path.join(UPLOAD_DIR, "bank-statements")
    os.makedirs(statement_dir, exist_ok=True)
    fname = f"{uuid.uuid4().hex}{ext}"
    full_path = os.path.join(statement_dir, fname)
    content = await file.read()
    with open(full_path, "wb") as fp:
        fp.write(content)

    parsed_lines = _parse_statement_file(full_path, currency.upper())
    if not parsed_lines:
        raise HTTPException(400, "No transaction lines found. Make sure the file has date and amount/debit/credit columns.")

    statement = BankStatementImport(
        bank_name=bank_name,
        account_name=account_name,
        account_no=account_no,
        original_filename=file.filename,
        file_path=f"accounting/bank-statements/{fname}",
        currency=currency.upper(),
        line_count=len(parsed_lines),
        uploaded_by_id=current_user.id,
        statement_from=min(line["transaction_date"] for line in parsed_lines),
        statement_to=max(line["transaction_date"] for line in parsed_lines),
    )
    db.add(statement)
    db.flush()

    lines: list[BankStatementLine] = []
    for line_data in parsed_lines:
        line = BankStatementLine(statement_id=statement.id, **line_data)
        db.add(line)
        lines.append(line)
    db.flush()
    for line in lines:
        _match_line(db, line)
    db.commit()
    db.refresh(statement)

    matched = sum(1 for line in lines if line.match_status == BankLineMatchStatus.MATCHED.value)
    possible = sum(1 for line in lines if line.match_status == BankLineMatchStatus.POSSIBLE.value)
    unmatched = sum(1 for line in lines if line.match_status == BankLineMatchStatus.UNMATCHED.value)
    return BankStatementUploadResponse(
        statement=statement,
        total_lines=len(lines),
        matched=matched,
        possible=possible,
        unmatched=unmatched,
    )


@router.get("/bank-statements/{statement_id}/lines", response_model=BankStatementLineListResponse)
def list_bank_statement_lines(
    statement_id: int,
    match_status: Optional[BankLineMatchStatus] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(BankStatementLine).filter(BankStatementLine.statement_id == statement_id)
    if match_status:
        q = q.filter(BankStatementLine.match_status == match_status.value)
    results = q.order_by(BankStatementLine.transaction_date.desc(), BankStatementLine.id.desc()).all()
    return BankStatementLineListResponse(total=len(results), results=results)


@router.post("/bank-statement-lines/{line_id}/confirm-match/{entry_id}", response_model=BankStatementLineListResponse)
def confirm_bank_line_match(
    line_id: int,
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.STAFF)),
):
    line = db.query(BankStatementLine).filter(BankStatementLine.id == line_id).first()
    if not line:
        raise HTTPException(404, "Bank statement line not found")
    entry = db.query(AccountingEntry).filter(AccountingEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Accounting entry not found")
    line.match_status = BankLineMatchStatus.MATCHED.value
    line.matched_entry_id = entry.id
    line.match_confidence = 100
    line.match_reason = "Confirmed manually"
    line.reviewed_by_id = current_user.id
    line.reviewed_at = datetime.utcnow()
    db.commit()
    results = db.query(BankStatementLine).filter(BankStatementLine.statement_id == line.statement_id).all()
    return BankStatementLineListResponse(total=len(results), results=results)


@router.post("/bank-statement-lines/{line_id}/unmatch", response_model=BankStatementLineListResponse)
def unmatch_bank_line(
    line_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.STAFF)),
):
    line = db.query(BankStatementLine).filter(BankStatementLine.id == line_id).first()
    if not line:
        raise HTTPException(404, "Bank statement line not found")
    line.match_status = BankLineMatchStatus.UNMATCHED.value
    line.matched_entry_id = None
    line.match_confidence = 0
    line.match_reason = "Unmatched manually"
    line.reviewed_by_id = current_user.id
    line.reviewed_at = datetime.utcnow()
    db.commit()
    results = db.query(BankStatementLine).filter(BankStatementLine.statement_id == line.statement_id).all()
    return BankStatementLineListResponse(total=len(results), results=results)


@router.get("/bank-statements/{statement_id}/download")
def download_bank_statement(
    statement_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    statement = db.query(BankStatementImport).filter(BankStatementImport.id == statement_id).first()
    if not statement:
        raise HTTPException(404, "Bank statement not found")
    uploads_root = os.path.dirname(UPLOAD_DIR)
    full_path = os.path.join(uploads_root, statement.file_path)
    if not os.path.isfile(full_path):
        raise HTTPException(404, "Bank statement file not found on disk")
    return FileResponse(full_path, filename=statement.original_filename)
